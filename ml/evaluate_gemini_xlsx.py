from __future__ import annotations

import argparse
from collections import Counter
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ml.signal_model import compute_signal


def infer_category(question: str, description: str) -> str:
    text = f"{question} {description}".lower()
    if any(k in text for k in ["stock", "shares", "earnings", "fed", "rates", "inflation", "crypto", "finance", "economic"]):
        return "financial"

    # Non-financial domains (politics, sports, elections, etc.) are treated as one bucket.
    return "general"


def parse_market_prob(last_trade_price: Any, outcome_prices: Any) -> float:
    if last_trade_price is not None:
        try:
            return float(last_trade_price)
        except (TypeError, ValueError):
            pass

    try:
        parsed = json.loads(str(outcome_prices))
        if isinstance(parsed, list) and parsed:
            return float(parsed[0])
    except (TypeError, ValueError, json.JSONDecodeError):
        pass

    return 0.5


def get_summary_input(row: list[Any], summary_idx: dict[str, int]) -> tuple[str, str]:
    """Pick pre-thesis summary text first, then fall back to thesis text."""
    for col_name in ("reasoning_input", "thesis_text"):
        col_idx = summary_idx.get(col_name)
        if col_idx is None:
            continue
        text = str(row[col_idx] or "").strip()
        if text:
            return text, col_name
    return "", ""


def evaluate_workbook(xlsx_path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    wb = load_workbook(xlsx_path, data_only=True)
    markets_ws = wb["polymarket_markets"]
    gemini_ws = wb["market_gemini_summaries"]

    market_headers = [markets_ws.cell(1, c).value for c in range(1, markets_ws.max_column + 1)]
    market_idx = {name: i for i, name in enumerate(market_headers)}

    summary_headers = [gemini_ws.cell(1, c).value for c in range(1, gemini_ws.max_column + 1)]
    summary_idx = {name: i for i, name in enumerate(summary_headers)}

    market_by_pid: dict[int, list[Any]] = {}
    for row_idx in range(2, markets_ws.max_row + 1):
        row = [markets_ws.cell(row_idx, c).value for c in range(1, markets_ws.max_column + 1)]
        pid = row[market_idx["polymarket_id"]]
        if pid is None:
            continue
        market_by_pid[int(pid)] = row

    records: list[dict[str, Any]] = []
    used_summary_source: Counter[str] = Counter()

    for row_idx in range(2, gemini_ws.max_row + 1):
        row = [gemini_ws.cell(row_idx, c).value for c in range(1, gemini_ws.max_column + 1)]
        pid_raw = row[summary_idx["polymarket_id"]]
        summary_text, summary_source = get_summary_input(row, summary_idx)

        if pid_raw is None or not summary_text:
            continue

        pid = int(pid_raw)
        market_row = market_by_pid.get(pid)
        if market_row is None:
            continue

        question = str(market_row[market_idx["question"]] or "")
        description = str(market_row[market_idx["description"]] or "")
        market_prob = parse_market_prob(
            market_row[market_idx["last_trade_price"]],
            market_row[market_idx["outcome_prices"]],
        )
        category = infer_category(question, description)

        signal = compute_signal(summary=summary_text, market_prob=market_prob, category=category)
        used_summary_source.update([summary_source])

        records.append(
            {
                "polymarket_id": pid,
                "question": question,
                "category": category,
                "summary_source": summary_source,
                "summary_preview": summary_text[:180].replace("\n", " "),
                "sentiment_score": signal["news_sentiment"],
                "market_crowd_score": signal["market_sentiment"],
                "divergence": signal["divergence"],
                "action": signal["action"],
            }
        )

    total = len(records)
    action_counts = Counter(x["action"] for x in records)
    category_counts = Counter(x["category"] for x in records)

    summary = {
        "rows_processed": total,
        "summary_input_column_priority": ["reasoning_input", "thesis_text"],
        "summary_source_counts": dict(used_summary_source),
        "category_counts": dict(category_counts),
        "action_counts": dict(action_counts),
    }

    return summary, records


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Feed XLSX summaries into signal model and export signal outputs."
    )
    parser.add_argument("--xlsx", default="data/sample_markets_ml.xlsx", help="Input XLSX path")
    parser.add_argument("--out-json", default="data/gemini_eval_results.json", help="Summary + records JSON output path")
    parser.add_argument("--out-csv", default="data/gemini_eval_records.csv", help="Flat record CSV output path")
    args = parser.parse_args()

    summary, records = evaluate_workbook(Path(args.xlsx))

    out_json = Path(args.out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps({"summary": summary, "records": records}, indent=2), encoding="utf-8")

    out_csv = Path(args.out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    write_csv(out_csv, records)

    print(json.dumps({"summary": summary, "out_json": str(out_json), "out_csv": str(out_csv)}, indent=2))


if __name__ == "__main__":
    main()
